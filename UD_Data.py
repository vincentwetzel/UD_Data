import os
import re
import shutil
import pytesseract
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from dateutil import parser

# Tesseract path
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# Config
SOURCE_DIR = "to_sort/"
SORTED_DIR = "sorted/"
PROCESSED_DIR = "processed/"
AUDIT_FILE = "uber_trip_audit.xlsx"

os.makedirs(SORTED_DIR, exist_ok=True)
os.makedirs(PROCESSED_DIR, exist_ok=True)


def generate_record_key(record):
    dt = record.get("datetime")
    dt_str = dt.strftime("%m/%d/%Y %I:%M %p") if dt else "N/A"
    start = record.get("start_address", "N/A").strip().lower()
    end = record.get("end_address", "N/A").strip().lower()
    return f"{dt_str}|{start}|{end}"


def find(pattern, text, default="N/A"):
    match = re.search(pattern, text)
    return match.group(1).strip() if match and match.lastindex else match.group(0).strip() if match else default


def extract_fields(text):
    date_match = re.search(r"([A-Za-z]{3,9} \d{1,2}, \d{4})", text)
    time_match = re.search(r"(\d{1,2}:\d{2}\s?[APMapm]{2})", text)
    if date_match and time_match:
        try:
            dt = datetime.strptime(f"{date_match.group(1)} {time_match.group(1)}", "%b %d, %Y %I:%M %p")
        except:
            dt = None
    else:
        dt = None

    upfront_match = re.search(r"\$([\d.,]+)\s*Upfront fare", text)
    upfront_earnings = upfront_match.group(1) if upfront_match else "N/A"

    dur_match = re.search(r"\b(\d{1,3})\s*min\s*(\d{1,3})\s*sec\b", text)
    duration = f"{dur_match.group(1)} minutes, {dur_match.group(2)} seconds" if dur_match else "N/A"

    dist_match = re.search(r"\b([\d]{1,3}\.\d{1,2})\s*(mi|km)\b", text)
    distance = f"{dist_match.group(1)} {dist_match.group(2)}" if dist_match else "N/A"

    points_match = re.search(r"\b([15l])\s*point[s]?\s*earned\b", text, re.IGNORECASE)
    raw = points_match.group(1) if points_match else "N/A"
    points = "1" if raw.lower() == "l" else raw if raw != "N/A" else "N/A"

    return {
        "datetime": dt,
        "trip_type": find(r"\bUberX\b|\bUberXL\b|\bUberBlack\b", text),
        "earnings": find(r"Your earnings\s*\$?([\d.,]+)", text),
        "upfront_earnings": upfront_earnings,
        "fare": find(r"Fare\s*\$?([\d.,]+)", text),
        "promotion": find(r"Promotion\s*\$?([\d.,]+)", text),
        "tip": find(r"Tip\s*\$?([\d.,]+)", text, default="$0.00"),
        "start_address": find(r"Esprit Dr,.*", text),
        "end_address": find(r"N Downwater St,.*", text),
        "points": points,
        "duration": duration,
        "distance": distance,
        "verified": find(r"Verified:\s*(TRUE|FALSE)", text, default="TRUE"),
        "discrepancy": find(r"Discrepancy Flag:\s*(TRUE|FALSE)", text, default="FALSE"),
        "text": text
    }


def extract_ocr_data(image_path):
    try:
        text = pytesseract.image_to_string(Image.open(image_path), lang='eng')
        print(f"\n📄 OCR text from {os.path.basename(image_path)}:\n{text}\n{'-' * 40}")
        fields = extract_fields(text)
        fields["filename"] = os.path.basename(image_path)
        return fields
    except Exception as e:
        print(f"❌ OCR error: {str(e)}")
        return {"error": str(e)}


def rename_file(dt, position="TOP"):
    if dt:
        date_str = f"{dt.month}.{dt.day}.{dt.year}"
        time_str = dt.strftime("%H-%M-%S")  # 24-hour format with hyphens
        return f"{date_str} {time_str}-{position}.jpg"
    else:
        return f"UnknownDate UnknownTime-{position}.jpg"


def get_sorted_path(dt, filename):
    if dt:
        year = str(dt.year)
        month = f"{dt.month} - {dt.strftime('%B')}"
        day = f"{dt.day:02d}"
        folder_path = os.path.join(SORTED_DIR, year, month, day)
    else:
        folder_path = os.path.join(SORTED_DIR, "UnknownDate")
    os.makedirs(folder_path, exist_ok=True)
    return os.path.join(folder_path, filename)


def parse_amount(value):
    try:
        return float(value.replace("$", "").replace(",", ""))
    except:
        return None


def is_same_trip(a, b):
    def match(val1, val2):
        return val1 == val2 and val1 != "N/A" and val2 != "N/A"

    def earnings_match(x, y):
        x_val = parse_amount(x.get("upfront_earnings", "N/A"))
        y_fare = parse_amount(y.get("fare", "N/A"))
        y_promo = parse_amount(y.get("promotion", "0.00"))
        return x_val is not None and y_fare is not None and y_promo is not None and abs(
            x_val - (y_fare + y_promo)) < 0.01

    address_match = match(a["start_address"], b["start_address"]) and match(a["end_address"], b["end_address"])
    if not address_match:
        return False

    if a["upfront_earnings"] == "N/A" or b["upfront_earnings"] == "N/A":
        return True

    return earnings_match(a, b) or earnings_match(b, a)


def merge_fields(primary, secondary):
    merged = primary.copy()
    for key in secondary:
        if merged.get(key, "N/A") == "N/A" and secondary[key] != "N/A":
            merged[key] = secondary[key]
    return merged


def duration_to_seconds(duration_str):
    match = re.search(r"(\d+)\s*minutes?,\s*(\d+)\s*seconds?", duration_str)
    if match:
        minutes = int(match.group(1))
        seconds = int(match.group(2))
        return minutes * 60 + seconds
    return ""


def distance_to_miles(distance_str):
    match = re.search(r"([\d]{1,3}\.\d{1,2})\s*mi", distance_str)
    if match:
        return float(match.group(1))
    return ""


def format_dollar(value):
    try:
        num = float(value.replace("$", "").replace(",", ""))
        return f"${num:,.2f}"
    except:
        return value


def log_to_excel(records):
    try:
        if os.path.exists(AUDIT_FILE):
            wb = load_workbook(AUDIT_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append([
                "Date/Time", "Trip Type", "Your Earnings", "Fare", "Promotion", "Tip",
                "Start Address", "End Address", "Points Earned", "Duration (seconds)", "Distance (miles)",
                "Discrepancy Flag", "Verified"
            ])

        # Build set of existing record keys
        existing_keys = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            dt_cell = row[0]
            if isinstance(dt_cell, datetime):
                dt_str = dt_cell.strftime("%m/%d/%Y %I:%M %p")
            else:
                dt_str = str(dt_cell) if dt_cell else "N/A"

            start_cell = row[6].strip().lower() if row[6] else "N/A"
            end_cell = row[7].strip().lower() if row[7] else "N/A"
            key = f"{dt_str}|{start_cell}|{end_cell}"
            existing_keys.add(key)

        new_entries = 0
        duplicates_skipped = 0

        for record in records:
            dt_str = record["datetime"].strftime("%m/%d/%Y %I:%M %p") if record["datetime"] else "N/A"
            key = generate_record_key(record)
            if key in existing_keys:
                print(f"⚠️ Skipped duplicate: {record['filename']}")
                duplicates_skipped += 1
                continue

            ws.append([
                dt_str,
                record.get("trip_type", ""),
                format_dollar(record.get("earnings", "")),
                format_dollar(record.get("fare", "")),
                format_dollar(record.get("promotion", "")),
                format_dollar(record.get("tip", "")),
                record.get("start_address", ""),
                record.get("end_address", ""),
                record.get("points", ""),
                duration_to_seconds(record.get("duration", "")),
                distance_to_miles(record.get("distance", "")),
                record.get("discrepancy", ""),
                record.get("verified", "")
            ])
            existing_keys.add(key)
            new_entries += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        wb.save(AUDIT_FILE)
        if new_entries > 0:
            print(f"📝 Audit log updated with {new_entries} new entr{'y' if new_entries == 1 else 'ies'}")
        else:
            print(
                f"📝 No new entries added. {duplicates_skipped} duplicate entr{'y' if duplicates_skipped == 1 else 'ies'} skipped.")
    except Exception as e:
        print(f"❌ Audit log error: {str(e)}")


def process_images():
    raw_data = []
    trip_index = 1
    audit_log = []

    print(f"\n📂 Files found in {SOURCE_DIR}:")
    for f in os.listdir(SOURCE_DIR):
        print(f"  - {f}")

    for filename in sorted(os.listdir(SOURCE_DIR)):
        if filename.lower().endswith((".jpg", ".jpeg", ".png")):
            print(f"\n🔍 Processing file: {filename}")
            path = os.path.join(SOURCE_DIR, filename)
            data = extract_ocr_data(path)
            raw_data.append(data)

    matched = set()
    for i, entry in enumerate(raw_data):
        if i in matched:
            continue
        for j in range(i + 1, len(raw_data)):
            if j in matched:
                continue
            if is_same_trip(entry, raw_data[j]):
                print(f"🔗 Matched: {entry['filename']} ↔ {raw_data[j]['filename']}")

                dt = entry["datetime"] if entry["datetime"] else raw_data[j]["datetime"]

                renamed_1 = rename_file(dt, "TOP")
                renamed_2 = rename_file(dt, "BOTTOM")

                original_path_1 = os.path.join(SOURCE_DIR, entry["filename"])
                original_path_2 = os.path.join(SOURCE_DIR, raw_data[j]["filename"])

                sorted_path_1 = get_sorted_path(dt, renamed_1)
                sorted_path_2 = get_sorted_path(dt, renamed_2)

                shutil.copy2(original_path_1, sorted_path_1)
                shutil.copy2(original_path_2, sorted_path_2)

                shutil.move(original_path_1, os.path.join(PROCESSED_DIR, entry["filename"]))
                shutil.move(original_path_2, os.path.join(PROCESSED_DIR, raw_data[j]["filename"]))

                merged = merge_fields(entry, raw_data[j])
                merged["filename"] = f"{renamed_1}, {renamed_2}"
                audit_log.append(merged)

                matched.update([i, j])
                trip_index += 1
                break

    skipped_count = 0
    for k, entry in enumerate(raw_data):
        if k not in matched:
            # Skip unmatched image — do not log or move
            skipped_count += 1
            print(f"⚠️ Skipped unmatched image: {entry['filename']} remains in to_sort/")

    log_to_excel(audit_log)
    print(f"\n✅ Processing complete. {len(audit_log)} entries logged. {skipped_count} unmatched images skipped.")


if __name__ == "__main__":
    process_images()
